import requests
import json
import pandas as pd
from collections import defaultdict, OrderedDict
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ====== CONFIG ======
API_TOKEN = "your_kobo_api_token_here"
FORM_UID  = "your_form_id_here"

ASSET_URL = f"https://collect.nrc.no/api/v2/assets/{FORM_UID}/?format=json"
HEADERS   = {"Authorization": f"Token {API_TOKEN}"}

OUT_XLSX = "kobo_import_template.xlsx"
DROPDOWN_ROWS = 1000

# Always include these extra columns at the end of Template
EXTRA_TEMPLATE_COLUMNS = ["Comments"]

# logical types to exclude from Template (cleaner sheet)
SKIP_LOGICAL_TYPES = {"note", "calculated"}

# ====== HELPERS ======
def normalize_label(lbl):
    if lbl is None:
        return ""
    if isinstance(lbl, str):
        return lbl.strip()
    if isinstance(lbl, dict):
        for k in ("English (en)", "en", "label", "English"):
            if k in lbl and lbl[k]:
                return str(lbl[k]).strip()
        for v in lbl.values():
            if v:
                return str(v).strip()
        return ""
    if isinstance(lbl, list):
        out = []
        for item in lbl:
            if isinstance(item, str) and item.strip():
                out.append(item.strip())
            elif isinstance(item, dict):
                out.append(normalize_label(item))
        return ", ".join([x for x in out if x])
    return str(lbl).strip()

def get_list_name(q):
    if q.get("select_from_list_name"):
        return q["select_from_list_name"].strip()
    t = (q.get("type") or "").strip()
    for prefix in ("select_one ", "select_multiple "):
        if t.startswith(prefix):
            return t[len(prefix):].strip()
    return None

def logical_type(q):
    t = (q.get("type") or "").strip()
    if t.startswith("select_one ") or q.get("select_from_list_name"):
        return "enum"
    if t.startswith("select_multiple "):
        return "enum[]"
    mapping = {
        "text": "string", "string": "string", "integer": "integer", "decimal": "number",
        "date": "date", "time": "time", "datetime": "datetime",
        "geopoint": "geopoint", "geotrace": "geotrace", "geoshape": "geoshape",
        "barcode": "string", "image": "binary", "audio": "binary", "video": "binary",
        "file": "binary", "acknowledge": "boolean", "calculate": "calculated", "note": "note",
        "rank": "rank", "url": "string", "range": "number",
    }
    return mapping.get(t, t or "unknown")

def is_structural(qtype):
    # skip only structure markers; keep notes/calculates out later
    return qtype in {"begin_group", "end_group", "begin_repeat", "end_repeat"}

def make_unique_header(name, path, used):
    """Disambiguate duplicate question names by appending __<path>."""
    if name not in used:
        return name
    suffix = path.replace("/", "__")
    candidate = f"{name}__{suffix}"
    i = 2
    while candidate in used:
        candidate = f"{name}__{suffix}__{i}"
        i += 1
    return candidate

# ====== FETCH SCHEMA ======
resp = requests.get(ASSET_URL, headers=HEADERS, timeout=60)
if resp.status_code != 200:
    print("❌ Error:", resp.status_code, resp.text)
    raise SystemExit()

asset = resp.json()

# Save full schema for reference
with open("form_schema.json", "w", encoding="utf-8") as f:
    json.dump(asset, f, indent=4, ensure_ascii=False)
print("✅ Schema downloaded. Open 'form_schema.json' to explore.")

survey  = (asset.get("content", {}) or {}).get("survey", []) or []
choices = (asset.get("content", {}) or {}).get("choices", []) or []

# list_name -> [{value, label}]
choices_by_list = defaultdict(list)
for c in choices:
    list_name = c.get("list_name")
    if not list_name:
        continue
    value = (c.get("name") or "").strip()
    label = normalize_label(c.get("label") or c.get("labels"))
    choices_by_list[list_name].append({"value": value, "label": label})

# Walk survey; gather all non-structural questions
path = []
fields = []

for q in survey:
    qtype = (q.get("type") or "").strip()

    if qtype in {"begin_group", "begin_repeat"}:
        grp = q.get("name") or normalize_label(q.get("label")) or ("group" if qtype == "begin_group" else "repeat")
        path.append(str(grp))
        continue
    if qtype in {"end_group", "end_repeat"}:
        if path:
            path.pop()
        continue
    if is_structural(qtype):
        continue

    name = (q.get("name") or "").strip()
    if not name:
        continue

    full = "/".join([*path, name]) if name else "/".join(path) or "(unnamed)"
    ltype = logical_type(q)
    list_name = get_list_name(q)

    enums = None
    enum_note = None
    if ltype in ("enum", "enum[]"):
        if list_name:
            if list_name in choices_by_list:
                enums = choices_by_list[list_name]
            else:
                enum_note = f"Choices not found for list '{list_name}'."
        else:
            enum_note = "No list name found for this select."

    fields.append({
        "path": full,
        "name": name,
        "label": normalize_label(q.get("label")),
        "kobo_type": qtype,
        "logical_type": ltype,
        "enum_list_name": list_name,
        "enums": enums,
        "enum_note": enum_note,
        "required": bool(q.get("required")),
    })

# ====== BUILD EXCEL TEMPLATE (ALL FIELD NAMES, EXCLUDING NOTES/CALCULATED) ======
# Template headers: start, end, then ALL question names (except note/calculated), then extras
final_cols = ["start", "end"]
used = set(final_cols)
header_meta = {}  # header -> field meta

for f in fields:
    if f["logical_type"] in SKIP_LOGICAL_TYPES:
        continue  # exclude notes and calculated fields
    nm, pth = f["name"], f["path"]
    col = make_unique_header(nm, pth, used)
    final_cols.append(col)
    used.add(col)
    header_meta[col] = f

for extra in EXTRA_TEMPLATE_COLUMNS:
    if extra not in used:
        final_cols.append(extra)
        used.add(extra)
        header_meta.setdefault(extra, None)

# 1) Template sheet with one blank data row
df_template = pd.DataFrame(columns=final_cols)
df_template.loc[0] = [""] * len(final_cols)

# 2) Choices sheet (codes) for ALL select fields; pad to equal length for pandas
choice_cols = OrderedDict()   # header -> list of codes
choice_len  = {}              # header -> real non-padded length

for col in final_cols:
    meta = header_meta.get(col)
    if not meta:
        continue
    if meta.get("logical_type") in ("enum", "enum[]") and meta.get("enums"):
        vals = [e["value"] for e in meta["enums"] if e.get("value")]
        choice_cols[col] = vals
        choice_len[col]  = len(vals)

if choice_cols:
    max_len = max(choice_len.values()) if choice_len else 0
    padded = {k: (v + [""] * (max_len - len(v))) for k, v in choice_cols.items()}
    df_choices = pd.DataFrame(padded)
else:
    df_choices = None

# 3) fields_catalog sheet (full reference)
# NEW/CHANGED: exclude logical_type == "note" from fields_catalog
catalog_fields = [f for f in fields if f["logical_type"] != "note"]  # <-- change
df_catalog = pd.DataFrame([{
    "name": f["name"],
    "label": f["label"],
    "path": f["path"],
    "logical_type": f["logical_type"],
    "required": f["required"],
    "enum_list_name": f["enum_list_name"],
    "enum_choices_codes": ", ".join([e["value"] for e in (f["enums"] or []) if e.get("value")]) if f.get("enums") else ""
} for f in catalog_fields])

# ====== 4) XML_Formula sheet ======
# NEW/CHANGED: exclude logical_type == "note" from XML_Formula
xml_rows = []
for f in catalog_fields:  # <-- use catalog_fields which already excludes "note"
    name = f["name"]
    path = f["path"]

    if not name or not path:
        continue

    open_tag = f"<{name}>"
    value_expr = f"{{row['{name}']}}"
    close_tag = f"</{name}>"

    xml_rows.append({
        "name": name,
        "path": path,
        "start_tag": open_tag,
        "value_expr": value_expr,
        "end_tag": close_tag,
    })

    # Build full nested XML block
    parts = path.split("/")
    if len(parts) == 1:
        # No nesting
        full_xml = f"<{name}>{value_expr}</{name}>"
    else:
        inner = f"<{name}>{value_expr}</{name}>"
        for grp in reversed(parts[:-1]):
            inner = f"<{grp}>\n{inner}\n</{grp}>"
        full_xml = inner

    xml_rows.append({
        "name": "",  # Keep empty to distinguish
        "path": "",
        "start_tag": "",
        "value_expr": full_xml,
        "end_tag": "",
    })

df_xml = pd.DataFrame(xml_rows, columns=["name", "path", "start_tag", "value_expr", "end_tag"])

# ====== WRITE EXCEL ======
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    df_template.to_excel(writer, index=False, sheet_name="template")
    if df_choices is not None:
        df_choices.to_excel(writer, index=False, sheet_name="choices")
    df_catalog.to_excel(writer, index=False, sheet_name="fields_catalog")
    df_xml.to_excel(writer, index=False, sheet_name="XML_Formula")

    wb = writer.book
    ws_template = wb["template"]
    ws_template.freeze_panes = "A2"  # keep headers visible

    # Dropdowns for select fields (by final header)
    if df_choices is not None and choice_cols:
        choice_headers = list(choice_cols.keys())
        for col_idx, header in enumerate(final_cols, start=1):
            if header not in choice_cols:
                continue
            ch_col_idx = choice_headers.index(header) + 1
            ch_col_letter = get_column_letter(ch_col_idx)
            last_row = choice_len[header] + 1  # +1 for header row on 'choices'
            formula = f"choices!${ch_col_letter}$2:${ch_col_letter}${last_row}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=True)
            tmpl_col_letter = get_column_letter(col_idx)
            dv_range = f"${tmpl_col_letter}$2:${tmpl_col_letter}${DROPDOWN_ROWS}"
            dv.add(dv_range)
            ws_template.add_data_validation(dv)

    # Header comments (Label / Path / Type / Required / Choice codes)
    for col_idx, header in enumerate(final_cols, start=1):
        meta = header_meta.get(header) or {}
        label = (meta.get("label") or "").strip()
        path_val = (meta.get("path") or "").strip()
        ltype = meta.get("logical_type") or ""
        req = "Yes" if meta.get("required") else "No"
        enum_codes = ""
        if meta.get("enums"):
            enum_codes = ", ".join([e["value"] for e in meta["enums"] if e.get("value")])

        comment_lines = []
        if header in ("start", "end"):
            comment_lines.append("System meta field (auto-filled at submit time if left blank).")
        elif not meta:
            comment_lines.append("Custom column (not in schema).")
        else:
            if label:
                comment_lines.append(f"Label: {label}")
            if path_val and path_val != meta.get("name"):
                comment_lines.append(f"Path: {path_val}")
            if ltype:
                comment_lines.append(f"Type: {ltype}")
            comment_lines.append(f"Required: {req}")
            if enum_codes:
                comment_lines.append(f"Choices (codes): {enum_codes}")
                if ltype == "enum[]":
                    comment_lines.append("For select_multiple, enter space-separated codes (e.g., a b c).")

        if comment_lines:
            ws_template.cell(row=1, column=col_idx).comment = Comment("\n".join(comment_lines), "Generator")

        # Column width hint
        width_hint = max(len(header), len(label)) + 2
        ws_template.column_dimensions[get_column_letter(col_idx)].width = min(max(12, width_hint), 48)

print(f"✅ Excel template written to '{OUT_XLSX}'")
print(f"   Template columns ({len(final_cols)}): {', '.join(final_cols[:12])}{' ...' if len(final_cols)>12 else ''}")
if df_choices is not None:
    print("   Added 'choices' sheet with select codes and dropdowns.")
print("   'fields_catalog' & 'XML_Formula' now exclude logical type 'note'.")
